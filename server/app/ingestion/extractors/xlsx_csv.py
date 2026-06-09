import io
from typing import Optional

import openpyxl
import pandas as pd

from ...api.errors import DocumentParseError
from ...schemas.discovery import ParsedDocument, TableBlock

_XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def extract_xlsx(
    content: bytes,
    filename: str,
    max_rows: Optional[int] = None,
) -> ParsedDocument:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise DocumentParseError(f"Failed to parse XLSX '{filename}': {exc}") from exc

    table_blocks: list[TableBlock] = []
    warnings: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = [
            [str(cell if cell is not None else "").strip() for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
        # Skip entirely empty sheets
        if not any(any(cell for cell in row) for row in all_rows):
            continue

        headers = all_rows[0] if all_rows else []
        rows = all_rows[1:] if len(all_rows) > 1 else []

        # Filter blank rows (all cells empty)
        rows = [r for r in rows if any(r)]

        if max_rows and len(rows) > max_rows:
            rows = rows[:max_rows]
            warnings.append(
                f"Sheet '{sheet_name}' truncated to {max_rows} rows "
                "(MAX_APPS_PER_RUN limit)"
            )

        table_blocks.append(
            TableBlock(sheet_name=sheet_name, headers=headers, rows=rows)
        )

    return ParsedDocument(
        filename=filename,
        content_type=_XLSX_CONTENT_TYPE,
        text_blocks=[],
        table_blocks=table_blocks,
        metadata={"sheet_count": len(wb.sheetnames)},
        warnings=warnings,
    )


def extract_csv(
    content: bytes,
    filename: str,
    max_rows: Optional[int] = None,
) -> ParsedDocument:
    warnings: list[str] = []

    try:
        df = pd.read_csv(io.BytesIO(content))
    except Exception as exc:
        raise DocumentParseError(f"Failed to parse CSV '{filename}': {exc}") from exc

    if max_rows and len(df) > max_rows:
        df = df.head(max_rows)
        warnings.append(
            f"CSV truncated to {max_rows} rows (MAX_APPS_PER_RUN limit)"
        )

    headers = df.columns.tolist()
    rows = df.fillna("").astype(str).values.tolist()
    sheet_name = filename.rsplit(".", 1)[0]  # stem of filename

    return ParsedDocument(
        filename=filename,
        content_type="text/csv",
        text_blocks=[],
        table_blocks=[
            TableBlock(sheet_name=sheet_name, headers=headers, rows=rows)
        ],
        metadata={"row_count": len(df), "column_count": len(df.columns)},
        warnings=warnings,
    )
